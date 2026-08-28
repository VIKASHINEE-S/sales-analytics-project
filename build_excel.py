import pandas as pd

from openpyxl import Workbook

from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    Reference
)

from openpyxl.styles import Font, Alignment


# ==========================================
# 1. READ CSV FILE
# ==========================================

df = pd.read_csv("sales_data.csv")

print("CSV Data:")
print(df.head())


# ==========================================
# 2. CREATE EXCEL WORKBOOK
# ==========================================

wb = Workbook()

ws = wb.active
ws.title = "Sales Data"


# ==========================================
# 3. WRITE SALES DATA TO EXCEL
# ==========================================

headers = list(df.columns)

for col_num, header in enumerate(headers, start=1):
    ws.cell(
        row=1,
        column=col_num,
        value=header
    )


for row_num, row_data in enumerate(
    df.itertuples(index=False),
    start=2
):

    for col_num, value in enumerate(
        row_data,
        start=1
    ):

        ws.cell(
            row=row_num,
            column=col_num,
            value=value
        )


# ==========================================
# 4. CREATE SUMMARY SHEET
# ==========================================

ws2 = wb.create_sheet("Summary")


# ==========================================
# 5. DASHBOARD TITLE - CENTER
# ==========================================

ws2.merge_cells("A1:T1")

ws2["A1"] = "SALES DASHBOARD"

ws2["A1"].font = Font(
    size=20,
    bold=True
)

ws2["A1"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

ws2.row_dimensions[1].height = 35


# ==========================================
# 6. KPI CARDS
# ==========================================

# Total Revenue
ws2["A3"] = "Total Revenue"
ws2["A4"] = "=SUM('Sales Data'!F:F)"


# Total Orders
ws2["C3"] = "Total Orders"
ws2["C4"] = "=COUNTA('Sales Data'!A:A)-1"


# Average Revenue
ws2["E3"] = "Average Revenue"
ws2["E4"] = "=AVERAGE('Sales Data'!F:F)"


# Highest Revenue
ws2["G3"] = "Highest Revenue"
ws2["G4"] = "=MAX('Sales Data'!F:F)"


# ==========================================
# 7. FORMAT KPI CARDS
# ==========================================

for cell in ["A3", "C3", "E3", "G3"]:

    ws2[cell].font = Font(
        bold=True,
        size=12
    )

    ws2[cell].alignment = Alignment(
        horizontal="center"
    )


for cell in ["A4", "C4", "E4", "G4"]:

    ws2[cell].font = Font(
        bold=True,
        size=14
    )

    ws2[cell].alignment = Alignment(
        horizontal="center"
    )


# ==========================================
# 8. REGION SUMMARY
# ==========================================

ws2["A7"] = "Region"
ws2["B7"] = "Total Revenue"


regions = df["Region"].unique()


for i, region in enumerate(
    regions,
    start=8
):

    ws2.cell(
        row=i,
        column=1,
        value=region
    )

    formula = (
        f'=SUMIF(\'Sales Data\'!C:C,'
        f'A{i},'
        f'\'Sales Data\'!F:F)'
    )

    ws2.cell(
        row=i,
        column=2,
        value=formula
    )


# ==========================================
# 9. PRODUCT SUMMARY
# ==========================================

ws2["D7"] = "Product"
ws2["E7"] = "Total Revenue"


products_list = df["Product"].unique()


for i, product in enumerate(
    products_list,
    start=8
):

    ws2.cell(
        row=i,
        column=4,
        value=product
    )

    formula = (
        f'=SUMIF(\'Sales Data\'!B:B,'
        f'D{i},'
        f'\'Sales Data\'!F:F)'
    )

    ws2.cell(
        row=i,
        column=5,
        value=formula
    )


# ==========================================
# 10. BAR CHART - REVENUE BY REGION
# ==========================================

chart1 = BarChart()

chart1.type = "col"

chart1.title = "Total Revenue by Region"

chart1.y_axis.title = "Revenue"

chart1.x_axis.title = "Region"


data1 = Reference(
    ws2,
    min_col=2,
    min_row=7,
    max_row=len(regions) + 7
)


categories1 = Reference(
    ws2,
    min_col=1,
    min_row=8,
    max_row=len(regions) + 7
)


chart1.add_data(
    data1,
    titles_from_data=True
)

chart1.set_categories(
    categories1
)


chart1.width = 12
chart1.height = 7


ws2.add_chart(
    chart1,
    "G7"
)


# ==========================================
# 11. BAR CHART - REVENUE BY PRODUCT
# ==========================================

chart2 = BarChart()

chart2.type = "col"

chart2.title = "Total Revenue by Product"

chart2.y_axis.title = "Revenue"

chart2.x_axis.title = "Product"


data2 = Reference(
    ws2,
    min_col=5,
    min_row=7,
    max_row=len(products_list) + 7
)


categories2 = Reference(
    ws2,
    min_col=4,
    min_row=8,
    max_row=len(products_list) + 7
)


chart2.add_data(
    data2,
    titles_from_data=True
)

chart2.set_categories(
    categories2
)


chart2.width = 12
chart2.height = 7


ws2.add_chart(
    chart2,
    "G22"
)


# ==========================================
# 12. LINE CHART - REVENUE TREND
# ==========================================

line = LineChart()

line.title = "Revenue Trend"

line.y_axis.title = "Revenue"

line.x_axis.title = "Order ID"


# Revenue is column F
data3 = Reference(
    ws,
    min_col=6,
    min_row=1,
    max_row=ws.max_row
)


# Order ID is column A
categories3 = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)


line.add_data(
    data3,
    titles_from_data=True
)

line.set_categories(
    categories3
)


line.width = 12
line.height = 7


ws2.add_chart(
    line,
    "G37"
)


# ==========================================
# 13. PIE CHART - REVENUE SHARE BY REGION
# ==========================================

pie = PieChart()

pie.title = "Revenue Share by Region"


# Region names
labels = Reference(
    ws2,
    min_col=1,
    min_row=8,
    max_row=len(regions) + 7
)


# Revenue values
data4 = Reference(
    ws2,
    min_col=2,
    min_row=7,
    max_row=len(regions) + 7
)


pie.add_data(
    data4,
    titles_from_data=True
)

pie.set_categories(
    labels
)


pie.width = 10
pie.height = 7


# Add Pie Chart
ws2.add_chart(
    pie,
    "T7"
)


# ==========================================
# 14. COLUMN WIDTH
# ==========================================

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 5
ws2.column_dimensions["G"].width = 18


# ==========================================
# 15. SAVE EXCEL FILE
# ==========================================

wb.save("Sales_Dashboard.xlsx")


print("================================")
print("Sales Dashboard created successfully!")
print("File: Sales_Dashboard.xlsx")
print("================================")